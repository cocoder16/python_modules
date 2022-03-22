from openpyxl import load_workbook, Workbook

def get_sheet_all_data(excel_path: str, sheet_name: str, remove_header: bool = True):
    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name]
    data_matrix = []
    i = 2 if remove_header else 1

    while True:
        j = 1
        row = []
        while True:
            value = worksheet.cell(row=i, column=j).value

            if value == None:
                break

            row.append(value)

            j += 1

        if row == []:
            break

        data_matrix.append(row)
        i += 1

    workbook.close()
    return data_matrix


def update_range(
    data_matrix: list,
    excel_path: str,
    sheet_name: str,
    row_start_index: int,
    column_start_index: int,
):
    if not data_matrix:
        return

    workbook = load_workbook(excel_path)
    # sheet_names = workbook.get_sheet_names()
    worksheet = workbook[sheet_name]

    row_length = len(data_matrix)
    column_length = len(data_matrix[0])
    
    for matrix_row_index in range(0, row_length):
        for matrix_column_index in range(0, column_length):
            worksheet.cell(
                row=matrix_row_index + row_start_index,
                column=matrix_column_index + column_start_index,
                value=data_matrix[matrix_row_index][matrix_column_index],
            )

    workbook.save(excel_path)
    workbook.close()


def delete_rows(excel_path: str, sheet_name: str, index: int = 1, amount: int = 1):
    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name]
    worksheet.delete_rows(index, amount)

    workbook.save(excel_path)
    workbook.close()


def create_excel(
    excel_path: str,
    sheet_name: str,
    initial_data: list = None,
    row_start_index: int = None,
    column_start_index: int = None,
):
    workbook = Workbook()
    workbook.create_sheet(title=sheet_name)
    worksheet = workbook[sheet_name]

    if workbook["Sheet"]:
        workbook.remove(workbook["Sheet"])

    if initial_data:
        row_length = len(initial_data)
        column_length = len(initial_data[0])
        for matrix_row_index in range(0, row_length):
            for matrix_column_index in range(0, column_length):
                worksheet.cell(
                    row=matrix_row_index + row_start_index,
                    column=matrix_column_index + column_start_index,
                    value=initial_data[matrix_row_index][matrix_column_index],
                )

    workbook.save(excel_path)
    workbook.close()


# header 포함한 총 row 개수
def get_row_size(excel_path: str, sheet_name: str):
    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name]

    return worksheet.max_row


def remove_duplicated_row_by_first_column(excel_path: str, sheet_name: str):
    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name]
    column_size = worksheet.max_column
    initial_row_length = worksheet.max_row - 1
    values = []
    data_matrix = []

    # reverse로 하는 이유는 더 마지막에 스크랩해온 row를 남기고 이전 row를 지우기 위해서이다.
    for i in range(initial_row_length + 1, 1, -1):
        if worksheet.cell(row=i, column=1).value in values:
            pass
        else:
            values.append(worksheet.cell(row=i, column=1).value)
            row = []

            for j in range(1, column_size + 1):
                value = worksheet.cell(row=i, column=j).value
                row.append(value)

            data_matrix.append(row)

    row_index = len(values) + 2
    amount = initial_row_length - len(values)
    worksheet.delete_rows(row_index, amount)
    workbook.save(excel_path)
    update_range(data_matrix, excel_path, sheet_name, 2, 1)
